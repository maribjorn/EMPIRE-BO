import pandas as pd
from openpyxl import load_workbook

# Path to your Excel file
file_path = 'Data handler/bolivia_v1/Node.xlsx'
workbook = load_workbook(file_path)

Horizon = 2040
start_year = 2020
LeapYearsInvestment = 5
new_names = ['Santivanez', 'San Julian', 'Las Carreras', 'Mazocruz', 'Litio/Lipez', 'Carrasco', 'Sucre', 'Yacuiba', 'San Borja', 'Vinto/Emvinto', 'Paraiso',
                 'Bélgica', 'Torre Huayco', 'Potosi 115', 'San Jose', 'Caranavi', 'Guarayos', 'Padilla', 'Choquetanga', 'Brechas', 'Bermejo', 'Catavi', 'Villazon',
                 'Yapacani', 'Chimoré', 'Punutuma', 'Los Troncos', 'Cumbre', 'San Buenaventura', 'Palca']  # Update this list as needed

total_periods = int((Horizon - start_year)/LeapYearsInvestment)
for sheet in workbook.sheetnames:
    current_sheet = workbook[sheet]

    nodes_positions = {'A3': ('B3', 1), 'E3': ('F3', 5)}
    
    for node_cell_ref, (period_cell_ref, col_index) in nodes_positions.items():
        if current_sheet[node_cell_ref].value == 'Nodes':
            # Check if the cell next to "Nodes" contains "Period"
            if current_sheet[period_cell_ref].value == 'Period':
                # Logic for handling periods
                start_row = current_sheet[node_cell_ref].row + 1
                
                # Repeat names for each period
                for period_index in range(total_periods):
                    period_start_row = start_row + (period_index * len(new_names))
                    
                    # Fill names for the current period
                    for i, name in enumerate(new_names, start=period_start_row):
                        current_sheet.cell(row=i, column=col_index, value=name)
                        current_sheet.cell(row=i, column=col_index + 1, value=period_index + 1)

                # Clear cells after the last period's names, if necessary
                last_name_row = start_row + (total_periods * len(new_names))
                for i in range(last_name_row, current_sheet.max_row + 1):
                    if current_sheet.cell(row=i, column=col_index).value or current_sheet.cell(row=i, column=col_index + 1).value:
                        current_sheet.cell(row=i, column=col_index).value = None
                        current_sheet.cell(row=i, column=col_index + 1).value = None
            else:
                max_row = current_sheet[node_cell_ref].row + len(new_names)
                # Fill names directly if no "Period" is next to "Nodes"
                for i, name in enumerate(new_names, start=current_sheet[node_cell_ref].row + 1):
                    current_sheet.cell(row=i, column=col_index, value=name)

                for i in range(max_row + 1, current_sheet.max_row + 1):
                    if current_sheet.cell(row=i, column=col_index).value is not None:
                        current_sheet.cell(row=i, column=col_index).value = None

workbook.save(file_path)


